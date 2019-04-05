# -*- coding: utf-8 -*-
"""
Created on Wed Jul 04 11:56:35 2012

@author: MosePetr
"""

# Read / Write XlSX Data
import openpyxl
import sqlite3
import io
import configparser
import wx

# import guidata.dataset.datatypes as dt
# import guidata.dataset.dataitems as di
import datetime, os


def CursorExecute(cursor, command, *args):
    """ 
    Execute SQL commmand, print debug info.
    """
    # print(command)
    n1 = datetime.datetime.now()
    cursor.execute(command, *args)
    n2 = datetime.datetime.now()
    print(command, args, (n2 - n1).microseconds, "\nRowCount", cursor.rowcount)


def insertIntoDb(filename, cursor, table, columns):
    """
    Insert data from Excel into DB 
    """
    # workbook = openpyxl.load_workbook(filename = filename, use_iterators = True)
    workbook = openpyxl.load_workbook(filename=filename)
    worksheet = workbook["List1"]
    print(table)
    sql = "insert into " + table + " values (? " + ", ?" * (columns - 1) + ")"
    # Insert data from ROW 2
    sqlrows = []
    firstrow = True
    for row in worksheet.iter_rows():
        values = []
        if firstrow:
            firstrow = False
        else:
            for column in range(0, columns):
                value = row[column].value
                if isinstance(value, datetime.datetime):
                    value = value.strftime("%Y%m%d")
                values.append(value)
            sqlrows.append(values)
    cursor.executemany(sql, sqlrows)
    print(sql)
    print(sqlrows[0:10])


def insertIntoXlsx(filename, cursor, table, columns):

    wb = openpyxl.Workbook()
    dest_filename = filename
    ws = wb.active
    ws.title = table
    print(table)

    sql = "select * from " + table
    row_idx = 1
    col_idx = 1
    for value in columns:
        ws.cell(column=col_idx, row=row_idx).value = value
        col_idx = col_idx + 1
    row_idx = row_idx + 1
    for row in cursor.execute(sql):
        col_idx = 1
        for value in row:
            ws.cell(column=col_idx, row=row_idx).value = value
            col_idx = col_idx + 1
        row_idx = row_idx + 1

    wb.save(filename=dest_filename)


class InsertFileNames(object):
    u"""Vstupni a vystupni soubory PRO1000"""

    def __init__(self):
        self.fileCelorocni = ""
        self.fileCji3 = ""
        self.fileCoois = ""
        self.fileCooisWzb = ""
        self.fileMb51 = ""
        self.fileStrojNastroj = ""
        self.filePrumery = ""
        self.fileToolbox = ""
        self.fileVystup = ""
        self.rok = 2019
        self.mesic = 3


# class InsertFileNames(dt.DataSet):
#     u"""Vstupni a vystupni soubory PRO1000"""
#     fileCelorocni = di.FileOpenItem("Celorocni", "xlsx")
#     fileCji3 = di.FileOpenItem("Cji3", "xlsx")
#     fileCoois = di.FileOpenItem("Coois", "xlsx")
#     fileCooisWzb = di.FileOpenItem("CooisWzb", "xlsx")
#     fileMb51 = di.FileOpenItem("Mb51", "xlsx")
#     fileStrojNastroj = di.FileOpenItem("StrojNastroj", "xlsx")
#     filePrumery = di.FileOpenItem("Prumery", "xlsx")
#     fileToolbox = di.FileOpenItem("Toolbox", "xlsx")
#     # fileSqlite3 = di.FileSaveItem("Sqlite3","db")
#     fileVystup = di.FileSaveItem("Vystup", "xlsx")
#     # fileCji3Vystup = di.FileSaveItem("Cji3Vystup","xlsx")
#     # fileMb51Vystup = di.FileSaveItem("Mb51Vystup","xlsx")
#     rok = di.IntItem("Rok vypoctu", min=2012, max=2022, default=2019)
#     mesic = di.IntItem("Mesic vypoctu", min=1, max=12, default=3)


def ConfigGet(config, section, option):
    if config.has_option(section, option):
        output = config.get(section, option)
    else:
        output = u""
    print(section, option, output)
    return output


def ConfigSet(config, section, option, value):
    config.set(section, option, value)
    print(section, option, value)


def UpdateNumeric(table, column):
    return (
        " UPDATE "
        + table
        + " set "
        + column
        + " = cast ("
        + column
        + " as Numeric) where "
        + column
        + " = cast ("
        + column
        + " as Numeric)"
    )


def CreateTables(cursor):
    tables = list(
        cursor.execute("select name from sqlite_master where type is 'table'")
    )
    for i in tables:
        CursorExecute(cursor, "drop table if exists %s" % i)
    # Import Excel Data into SQLite.
    CursorExecute(
        cursor,
        """create table celorocni (Zakazka, Nastroj, Stroj, 
        Stredisko, NaklStredisko, Popis, Typ, Skupina, Graf)""",
    )
    CursorExecute(
        cursor,
        """create table cji3 (Objekt, Objektbezeichnung, RefBelegnr, 
        Kostenart, Benutzername, WertKWahr, KWahr, Menge, GME, BuchDatum, ErfasstAm)""",
    )
    CursorExecute(
        cursor,
        "create table cooispro1000 (Auftrag, LeistArt, Vorgang, ArbPlatz, Kurztext)",
    )
    CursorExecute(
        cursor,
        """create table cooiswzb (Auftrag,Material, Ikone, AufArt, 
        Disponent, FertSteu, Werk, Typ, Systemstatus, Version, Materialkurztext, 
        Sollmenge, Einheit, Eckstart, Eckende)""",
    )

    CursorExecute(
        cursor,
        """ create table mb51 (Bukr, Material, Materialkurztext, 
        Werk, LOrt, BwA, Bewegungsartentext, S, MatBeleg, Pos, Kunde, Charge, 
        Benutzer, Uhrzeit, Bestellung, Auftrag, Text, Kostenst,Buchdat, 
        MengeInEME, EME, BetragHauswahr, BPM, Verkaufswert)""",
    )

    # CursorExecute(cursor, """ create table toolbox (Kostenart,KstArtBez, Objekt,
    #   Objektbezeichnung, Per, Benutzername, Material, Materialkurztext, WertBW,
    #    Mengeerf, GME,Belegdatum)""")
    CursorExecute(
        cursor,
        """ 
        create table toolbox (Benutzer, Material, Materialkurztext, Kostenstelle,
          Text, Hauswaehr, MngEME, EME,
          Belegdatum, Waehrg)""",
    )

    CursorExecute(cursor, "create table strojnastroj (Nastroj, Stroj)")
    CursorExecute(cursor, "create table prumery (Graf, Prumer, Cil)")


def AlterToolbox(localcursor):
    CursorExecute(localcursor, "alter table toolbox add column Auftrag9")
    CursorExecute(localcursor, "alter table toolbox add column Jahr")
    CursorExecute(localcursor, "alter table toolbox add column Monat")
    CursorExecute(localcursor, "alter table toolbox add column Skupina")
    CursorExecute(localcursor, "alter table toolbox add column Nastroj")
    CursorExecute(localcursor, "alter table toolbox add column PracMisto")

    CursorExecute(
        localcursor,
        """
        update  toolbox set Auftrag9 = (select max(Zakazka) from celorocni
            where ((KostenStelle = NaklStredisko) or
        (cast(KostenStelle as integer) = cast(NaklStredisko as integer))))""",
    )
    CursorExecute(
        localcursor,
        """
        update  toolbox set Skupina = (select max(Skupina) from celorocni
        where ((KostenStelle = NaklStredisko)
        or (cast(KostenStelle as integer)= cast(NaklStredisko as integer))))
        """,
    )
    CursorExecute(
        localcursor,
        """update  toolbox set Skupina = NULL where ((Skupina = 0) or (cast(Skupina as integer)= 0))""",
    )
    CursorExecute(
        localcursor,
        """        
        update  toolbox set Nastroj = (select max(Nastroj) from celorocni 
        where ((KostenStelle = NaklStredisko) 
        or (cast(KostenStelle as integer)= cast(NaklStredisko as integer))))""",
    )
    CursorExecute(
        localcursor,
        """        
        update  toolbox set PracMisto = (select max(Stroj) from celorocni 
        where ((KostenStelle = NaklStredisko) 
        or (cast(KostenStelle as integer)= cast(NaklStredisko as integer))))""",
    )
    CursorExecute(
        localcursor, """ update  toolbox set Jahr = substr(Belegdatum,1,4) """
    )
    CursorExecute(
        localcursor, """ update  toolbox set Monat = substr(Belegdatum,5,2) """
    )
    # set PracMisto, Skupina, Nastroj according to Auftrag9
    CursorExecute(
        localcursor,
        """ 
        update toolbox set PracMisto = 
            (select max(Stroj) from celorocni where Auftrag9 = Zakazka)
            where PracMisto is Null
        """,
    )
    CursorExecute(
        localcursor,
        """ 
        update toolbox set Nastroj = 
            (select max(Nastroj) from celorocni where Auftrag9 = Zakazka)
            where Nastroj is Null
        """,
    )
    # Kurs CZK/EUR
    CursorExecute(
        localcursor,
        """ 
                  update Toolbox set Hauswaehr = Hauswaehr / -26
        """,
    )


def LoadConfig():
    pass


def w32FillExcel():
    from win32com.client import Dispatch

    # Open Excel workbook
    xl = Dispatch("Excel.Application")
    wbinput = xl.Workbooks.Open(prm.fileVystup)
    wsinput = wbinput.Worksheets("spojenyvystup")
    for stredisko in [
        "automaty",
        "poloautomaty",
        "rucni",
        "stanzen",
        "trennen",
        "vstrikovna",
    ]:
        wb = xl.Workbooks.Open(path + "\\" + "naklady-" + stredisko + ".xlsx")
        ws = wb.Worksheets("spojenyvystup")
        wsinput.Range("A:T").Copy(ws.Range("A1"))
        wb.Worksheets(2).Activate()
        wb.RefreshAll()
        wb.Save()
        wb.Close()
    wbinput.Close()
    xl.Quit()


class GuiReadFiles(wx.Panel):
    def __init__(self, parent):
        super().__init__(parent)
        button_ok = wx.Button(self, label="OK")
        button_ok.Bind(wx.EVT_BUTTON, parent.on_close)


class GuiFrame(wx.Frame):
    def __init__(self):
        super().__init__(None, title="Zadej soubory")
        """
        Create a toolbar
        """
        self.toolbar = self.CreateToolBar()
        self.toolbar.SetToolBitmapSize((16, 16))
        self.folderPath = ""
        open_ico = wx.ArtProvider.GetBitmap(wx.ART_FILE_OPEN, wx.ART_TOOLBAR, (16, 16))
        OpenTool = self.toolbar.AddTool(
            wx.ID_ANY, "Otevřít", open_ico, "Otevřít složku"
        )
        self.Bind(wx.EVT_MENU, self.on_open_directory, OpenTool)

        panel = GuiReadFiles(self)
        self.Show()

    def on_open_directory(self, event):
        """
        Open a directory dialog
        :param event:
        :return:
        """
        with wx.DirDialog(self, "Vyberte složku", style=wx.DD_DEFAULT_STYLE) as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                self.folderPath = dlg.GetPath()

    def on_close(self, event):
        self.Close(force=False)


def run_wx():
    app = wx.App(redirect=False)
    frame = GuiFrame()
    app.MainLoop()


if __name__ == "__main__":
    # Create QApplication
    # import guidata

    # _app = guidata.qapplication()
    run_wx()
    prm = InsertFileNames()
    config = configparser.RawConfigParser(allow_no_value=True)
    config.read("pro1000.cfg")
    for (param, value) in vars(prm):
        if isinstance(object, di.StringItem):
            filename = ConfigGet(config, prm.__doc__, object._name)
            prm.__setattr__(object._name, filename)
            print(prm.__getattribute__(object._name))
    # prm.edit()
    # prm.view()

    # config = configparser.RawConfigParser()
    # config.add_section(prm.__doc__)
    # for object in prm._items:
    #     if isinstance(object, di.StringItem) or isinstance(object, di.IntItem):
    #         ConfigSet(config, prm.__doc__, object._name, object.get_string_value(prm))
    #
    # with io.open("pro1000.cfg", "w", encoding="utf-8") as configfile:
    #     config.write(configfile)

    # Open SQLite Database.
    ###
    ### Opening/creating database
    ###

    # connection=sqlite3.Connection(prm.fileSqlite3)
    file = prm.fileVystup
    path = os.path.dirname(file)
    fileSqlite3db = path + "\\Sqlite3.db"
    # connection=sqlite3.Connection(":memory:")
    try:
        os.remove(fileSqlite3db)
    except OSError as err:
        print(u"File " + fileSqlite3db + u" didn't exist")
    connection = sqlite3.Connection(fileSqlite3db)
    # connection=sqlite3.Connection(":memory:")
    cursor = connection.cursor()
    CreateTables(cursor)

    insertIntoDb(prm.fileCelorocni, cursor, "celorocni", 9)
    insertIntoDb(prm.fileCji3, cursor, "cji3", 11)
    insertIntoDb(prm.fileCoois, cursor, "cooispro1000", 5)
    insertIntoDb(prm.fileCooisWzb, cursor, "cooiswzb", 15)
    insertIntoDb(prm.fileMb51, cursor, "mb51", 24)
    insertIntoDb(prm.fileStrojNastroj, cursor, "strojnastroj", 2)
    insertIntoDb(prm.filePrumery, cursor, "prumery", 3)
    insertIntoDb(prm.fileToolbox, cursor, "toolbox", 10)

    # Convert ArbPlatz, Wkz to numbers where approtiate.
    CursorExecute(cursor, UpdateNumeric("cooispro1000", "ArbPlatz"))
    CursorExecute(cursor, UpdateNumeric("celorocni", "nastroj"))
    CursorExecute(cursor, UpdateNumeric("celorocni", "stroj"))

    # Now add necessary fields
    # Delete from table cooispro1000 where ArbPlatz in ( 6666, 7777)
    CursorExecute(
        cursor,
        """ Delete from cooispro1000 where ArbPlatz in ( 6666, 7777, 1006666, 1007777)""",
    )
    CursorExecute(
        cursor,
        """ Delete from cooispro1000 where ArbPlatz in ( '6666', '7777', '1006666', '1007777')""",
    )

    CursorExecute(
        cursor,
        """ Create table cooisnastroj as select * from cooispro1000 where LeistArt = 'WKZ' """,
    )
    CursorExecute(cursor, """ Create index cooisnastroj_i on cooisnastroj (auftrag) """)

    CursorExecute(cursor, """ Delete from cooispro1000 where LeistArt = 'WKZ' """)
    CursorExecute(
        cursor,
        """ create table cooisvystup as select distinct c.*,n.ArbPlatz as Wkz  
                       from cooispro1000 as c left outer join  cooisnastroj as n on c.Auftrag = n.Auftrag """,
    )

    CursorExecute(cursor, """ create index cooisvystup_i on cooisvystup (auftrag) """)
    CursorExecute(
        cursor, """ create index cooisvystup_av on cooisvystup (auftrag,vorgang) """
    )
    CursorExecute(
        cursor, """ update cooisvystup set wkz = ArbPlatz where Wkz is null """
    )
    #    CursorExecute(cursor, """ delete from cooisvystup where vorgang != 0130 """)
    # MB51
    #    CursorExecute(cursor, """ create table mb51vystup as
    #     select m.Material, m.MaterialKurztext, m.Uhrzeit, m.Buchdat, substr(m.Buchdat,1,4) as Jahr,  substr(m.Buchdat,5,2) as Monat,
    #        m.Auftrag, m.MengeInEME, c.ArbPlatz, c.Wkz
    #        from mb51 as m, cooisvystup as c where c.auftrag = m.auftrag """)
    #
    CursorExecute(
        cursor,
        """ create table mb51vystup as      
     select m.Material, m.MaterialKurztext, m.Uhrzeit, m.Buchdat, substr(m.Buchdat,1,4) as Jahr,  substr(m.Buchdat,5,2) as Monat,
        m.Auftrag, m.MengeInEME, c.ArbPlatz, c.Wkz
        from mb51 as m, cooisvystup as c where c.auftrag = m.auftrag and (c.vorgang = '0130') """,
    )

    CursorExecute(
        cursor,
        """ create table mb51sumvystup as
        select material, Jahr, Monat, Auftrag, sum(MengeInEME)as Menge,  ArbPlatz, Wkz from mb51vystup
            group by material, Jahr, Monat, Auftrag, ArbPlatz, Wkz""",
    )

    CursorExecute(
        cursor,
        """        
        alter table mb51sumvystup add column Auftrag9""",
    )

    CursorExecute(
        cursor,
        """        
        update  mb51sumvystup set Auftrag9 = (select max(Zakazka) from celorocni 
            where ((nastroj = Wkz) or (cast(nastroj as integer)= cast(Wkz as integer))))""",
    )

    CursorExecute(
        cursor,
        """        
        update  mb51sumvystup set Auftrag9 = (select Zakazka from celorocni 
                where (cast(stroj as integer)= cast(Arbplatz as integer)) or (stroj = Arbplatz))
            where (Auftrag9 is Null) or (Auftrag9 = 0)
            """,
    )

    AlterToolbox(cursor)

    CursorExecute(
        cursor,
        """ delete from cji3 where (Objekt like '2%') and (Objekt not in (select Auftrag from cooiswzb))""",
    )
    CursorExecute(cursor, """ delete from cji3 where (Objekt like '2155105%') """)
    CursorExecute(
        cursor,
        """ create table cji3vystup as
    select c.Objekt as Objekt, c.Objektbezeichnung as Bezeichnung, sum(c.WertKWahr) as WertK, 
        substr(c.buchdatum,1,4) as Jahr ,substr(c.buchdatum,5,2) as Monat , w.material as Material, 
        w.materialkurztext as MaterialKurzText, substr(w.Material, 1, 4) as Wkz, substr(w.Material, 1, 4) as ArbPlatz, c.Objekt as Auftrag9
       from cji3 as c left outer join   cooiswzb as w on c.objekt = w.auftrag
       group by c.Objekt, substr(c.buchdatum,1,6)
    """,
    )

    CursorExecute(
        cursor,
        """ update  cji3vystup set Wkz= 
        (select max(Nastroj) from StrojNastroj where  (cast(Wkz as Integer)  = cast (Nastroj as Integer)) or (Wkz = Nastroj)) 
        where Objekt like '2%'
    """,
    )

    CursorExecute(
        cursor,
        """ update  cji3vystup set ArbPlatz= 
        (select max(Stroj) from StrojNastroj where  (cast(Wkz as Integer)  = cast (Nastroj as Integer)) or (Wkz = Nastroj)) 
        where Objekt like '2%'
    """,
    )

    CursorExecute(
        cursor,
        """ update  cji3vystup set Auftrag9= 
        (select max(Zakazka) from celorocni where  cast(Wkz as Integer)  = cast (Nastroj as Integer)) 
        where Objekt like '2%' 
    """,
    )

    # Sjednoceny vystup MB51sumvystup a CJI3Vystup
    CursorExecute(
        cursor,
        """ 
        create table spojenyvystup (Zakazka, Material, Text, Mnozstvi, NakladyEur, 
                             Nastroj, PracMisto, Zakazka9, Stredisko, Skupina,
                             Rok, Mesic , Graf, Cil, Prumer, KumulaceMnozstvi, KumulaceNakladyEur, KumulacePodil,Zdroj)
    """,
    )

    CursorExecute(
        cursor,
        """ 
         Insert into spojenyvystup select Auftrag, Material, "", Menge, 0,
            Wkz, ArbPlatz, Auftrag9, NaklStredisko, Skupina,
            cast(Jahr as Interger), cast(Monat as Integer), 0, 0, 0, 0, 0, "",""
            from mb51sumvystup left outer join celorocni on Auftrag9 = Zakazka
        """,
    )

    CursorExecute(
        cursor,
        """ 
         Insert into spojenyvystup select Objekt, Bezeichnung, MaterialKurzText, 0, WertK,
            Wkz, ArbPlatz, Auftrag9, NaklStredisko, Skupina,
            cast(Jahr as Integer), cast(Monat as Integer), 0, 0, 0, 0, 0, "", ""
            from cji3vystup left outer join celorocni on cast(Auftrag9 as integer) = celorocni.Zakazka
        """,
    )

    CursorExecute(
        cursor,
        """
        Insert into spojenyvystup 
        (Zakazka, Material, Text, Mnozstvi,
         NakladyEur, Nastroj, PracMisto, Zakazka9,
         Stredisko, Skupina, Rok, Mesic,
         Zdroj)
        select Auftrag9, Material, Materialkurztext, 0,
         HausWaehr, Nastroj, PracMisto, Auftrag9,
         KostenStelle, Skupina, cast(Jahr as Integer), cast(Monat as Integer),
         'Toolbox'
         from toolbox 
        """,
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set pracMisto = (select max(Stroj)  from celorocni
        where (cast(zakazka9 as integer) = cast(zakazka as integer)) )
        where pracmisto is null""",
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set Nastroj = (select max(Nastroj)  from celorocni
        where (cast(zakazka9 as integer) = cast(zakazka as integer)) )
        where Nastroj is null
    """,
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set skupina = Nastroj where ((skupina is null) or skupina="" or skupina=3604) and 
        ((Stredisko in (40304102, 40350100, 40354101, 40354102 )) or graf in ("trennen","vstrikovna")) and (Nastroj > 0)""",
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set skupina = pracMisto where skupina is null and 
         graf in ("stanzen","rucni","poloautomaty","automaty")""",
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set skupina = pracMisto where skupina is null """,
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set Graf = (select max(Graf)  from celorocni
        where (cast(zakazka9 as integer) = cast(zakazka as integer)) )
        """,
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set skupina = Nastroj where 
        graf in ("trennen", "vstrikovna") and (length(trim(Nastroj))>0 )""",
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set skupina = Nastroj where (skupina is null or length(trim(skupina))=0) and 
        graf in ("trennen","vstrikovna") and (length(trim(Nastroj))>0 )""",
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set skupina = pracMisto where (skupina is null or length(trim(skupina))=0) and 
        graf in ("stanzen","rucni","poloautomaty","automaty") and (length(trim(pracMisto))>0)""",
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set skupina = 
        case when length(trim(pracMisto)) > 0 then pracMisto 
             when length(trim(Nastroj)) > 0 then Nastroj else "x" end 
        where skupina is null or length(trim(skupina)) =0 """,
    )

    CursorExecute(
        cursor,
        """ 
    update  spojenyvystup set zdroj = "Instandhaltungsauftrag"
        where zdroj is null or length(trim(zdroj)) =0 """,
    )

    CursorExecute(
        cursor,
        """delete from spojenyvystup where mesic > ? and rok=?""",
        (prm.mesic, prm.rok),
    )

    # Create rows for AVG ... Row function from table "Prumery"
    for t in range(1, 13):
        CursorExecute(
            cursor,
            """insert into spojenyvystup (Mesic, Rok, Graf, Cil, Prumer) 
            select ?, ?, Graf, Cil, Prumer from Prumery""",
            (t, prm.rok),
        )

    CursorExecute(cursor, """delete from spojenyvystup where Rok<>?""", (prm.rok,))

    CursorExecute(
        cursor,
        """insert into spojenyvystup (Mesic, Rok, Graf, Cil, Prumer, Mnozstvi, NakladyEur) 
            select ' Prumer 2019', 2019, Graf, Cil, Prumer, 1000, Prumer from Prumery """,
    )

    for t in range(1, prm.mesic + 1):
        CursorExecute(
            cursor,
            """insert into spojenyvystup (Mesic, Rok, Graf, KumulaceMnozstvi, KumulaceNakladyEur, KumulacePodil) 
            select ?, ?, Graf, sum(Mnozstvi), sum(NakladyEur), (sum(NakladyEur)/sum(Mnozstvi))*1000 from spojenyvystup where Mesic<=? group by  Graf""",
            (t, prm.rok, t),
        )

    insertIntoXlsx(
        prm.fileVystup,
        cursor,
        "spojenyvystup",
        "Zakazka, Material , Text, Mnozstvi, NakladyEur, Nastroj, PracMisto, Zakazka9, Stredisko, Skupina, Rok, Mesic, Graf, Cil, Prumer, KumulaceMnozstvi, KumulaceNakladyEur, KumulacePodil, Zdroj".split(
            ","
        ),
    )

    file = prm.fileVystup
    path = os.path.dirname(file)

    fileSqlite3 = path + "\\Sqlite3.sql"
    with io.open(fileSqlite3, "w", encoding="utf-8") as f:
        for line in connection.iterdump():
            f.write(u"%s\n" % line)
    try:
        CursorExecute(cursor, """commit""")
    except:
        pass
    connection.close()

    w32FillExcel()
